from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from util_funcs import get_cell
import os


def main():

    solution_wb = load_workbook('quizes/solution.xlsx', data_only=True)
    solution_ws = solution_wb.active
    results = dict()

    for file in os.listdir('quizes'):
        if file != 'solution.xlsx':
            wb = load_workbook(f'quizes/{file}', data_only=True)
            ws = wb.active
    
            name = file.split('.')[0]

            results[name] = compare_tip(ws, solution_ws)

    print(results)

    with open('result.txt', 'w') as result_file:
        for key, value in results.items():
            result_file.write(f'{key}: {value}\n')


def compare_tip(ws, solution_ws):

    total_points = 0
    # one point per correct scored goals and one point if correct match result
    total_points += group_points(ws, solution_ws, [3, 4, 5])
    # one point per correct scored goals and two points per correct team into round of 16 
    total_points += round_of_16_points(ws, solution_ws, 8, 2, [14, 15, 16, 17]) 
    # one point per correct scored goals and four points per correct team into quarterfinals
    total_points += quarter_points(ws, solution_ws, 4, 4, [19, 20, 21, 22]) 
    # one point per correct scored goals and six points per correct team into semifinals
    total_points += semi_points(ws, solution_ws, 2, 6, [24, 25, 26, 27])

    return total_points


def group_points(ws, solution_ws, col_range):
    points = 0
    row = 3
    last_match_row = 64
    for i in range(3, last_match_row+1):
        if row in [9, 17, 25, 33, 41, 49, 57]:
            row += 2
        else:
            for col in col_range: # C, D, E
                cell = get_cell(col, row)
                if ws[cell].value == solution_ws[cell].value:
                    points += 1
            row += 1
        if row > last_match_row:
            return points

# get teams from solution worksheet from playoff
def get_teams(solution_ws, n_matches, col_range):
    teams = []
    for row in range(4, n_matches*4+1, 4):
        for col in col_range:
            cell = get_cell(col, row)
            teams.append(solution_ws[cell].value)
    return teams

# helper function to get points from different parts in the playoff
def get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range):
    teams = get_teams(solution_ws, n_matches, col_range[0:3])
    points = 0
    for row in range(4, n_matches*4+1, 4):
        for col in col_range[0:2]:
            cell = get_cell(col, row)
            if ws[cell].value in teams:
                points += team_points
        for col in col_range[2:4]:
            cell = get_cell(col, row)
            if ws[cell].value == solution_ws[cell].value:
                points += 1
    return points


def round_of_16_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)
    

def quarter_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)


def semi_points(ws, solution_ws, n_matches, team_points, col_range):
    return get_playoffs_points(ws, solution_ws, n_matches, team_points, col_range)


# def final_points():
#     return get_playoffs_points(ws, solution_ws, n_matches, team_points)
    

if __name__ == '__main__':
    main()